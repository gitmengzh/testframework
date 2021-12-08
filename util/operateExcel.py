#   封装读取excel

"""
用xlrd和xlwt进行excel读写

用openpyxl进行excel读写；

用pandas进行excel读写；

"""
import openpyxl

def readExel(filename):
    if not filename:
        filename = r'D:\work\Excel_txtProcesss\new-微博-合并\58.xlsx'
    excel = openpyxl.load_workbook(filename)  # 读文件

    sheetnames = excel.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = excel.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    for r in range(1, rows):
        for c in range(1, cols):
            print(ws.cell(r, c).value)
        if r == 10:
            break


def writeExcel(self):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for row in range(1, 70000):
        for col in range(1, 4):
            outws.cell(row, col).value = row * 2  # 写文件
        print(row)
    saveExcel = "D:\\work\\Excel_txtProcesss\\test.xlsx"
    outwb.save(saveExcel)  # 一定要记得保存


# from packageRequests import basePost
# from packageRequests import baseGet
#
# res = baseGet("www.baidu.com")
# print(res)
#
#
# import  pysnooper
#
# @pysnooper.snoop()
# def number_to_bits(number):
#     if number == 1:
#         return 1
#     else:
#         return number
#
# number_to_bits(6)